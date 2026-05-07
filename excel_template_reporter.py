# Databricks notebook source
# MAGIC %md
# MAGIC ## XLSX Template Reporter - Usage Guide
# MAGIC
# MAGIC ### Overview
# MAGIC This notebook fills an XLSX template with SQL query results and free-form text,
# MAGIC preserving all formatting (fonts, colors, borders, number formats).
# MAGIC
# MAGIC ### Template Marker Convention
# MAGIC
# MAGIC Place markers in template cells using `{{name}}` syntax:
# MAGIC
# MAGIC | Marker Type | Example | Description |
# MAGIC |-------------|---------|-------------|
# MAGIC | Text | `{{report_title}}` | Replaced with the value of the matching widget |
# MAGIC | Text (inline) | `Generated: {{generated_date}}` | Marker portion replaced, surrounding text kept |
# MAGIC | Table | `{{table:sales}}` | Data rows inserted here from `source_sales` widget query |
# MAGIC
# MAGIC ### Table Marker Layout
# MAGIC
# MAGIC For a `{{table:sales}}` marker, the template must have this row structure:
# MAGIC - **Row above marker**: Header row (preserved as-is)
# MAGIC - **Marker row**: Template data row — formatting cloned for each result row
# MAGIC - **Row below marker**: Footer row — shifted down, formulas adjusted to span data
# MAGIC
# MAGIC ### Parameters (Widgets)
# MAGIC
# MAGIC | Parameter | Type | Description |
# MAGIC |-----------|------|-------------|
# MAGIC | `template_path` | Text | UC Volume path to template XLSX |
# MAGIC | `output_path` | Text | UC Volume directory for output |
# MAGIC | `output_filename` | Text | Output file name (default: report.xlsx) |
# MAGIC | `source_sales` | Text | SQL query or table name for `{{table:sales}}` |
# MAGIC | `report_title` | Text | Free-form text for `{{report_title}}` marker |
# MAGIC | `description` | Text | Free-form text for `{{description}}` marker |
# MAGIC | `generated_date` | Text | Free-form text for `{{generated_date}}` marker |
# MAGIC
# MAGIC Add one `source_<name>` widget per table marker in your template.
# MAGIC
# MAGIC ### Example Usage
# MAGIC
# MAGIC ```
# MAGIC template_path:  /Volumes/main/default/templates/sales_template.xlsx
# MAGIC output_path:    /Volumes/main/default/reports/
# MAGIC output_filename: q1_sales_report.xlsx
# MAGIC source_sales:   SELECT product, region, quantity, unit_price, total FROM main.default.sales WHERE quarter = 'Q1'
# MAGIC report_title:   Quarterly Sales Report - Q1 2026
# MAGIC description:    Summary of product sales across all regions.
# MAGIC generated_date: 2026-04-05
# MAGIC ```
# MAGIC
# MAGIC ### Notes
# MAGIC
# MAGIC * Data is collected to the driver — be mindful of dataset size
# MAGIC * The template XLSX must exist at the specified UC Volume path
# MAGIC * Output file overwrites any existing file with the same name
# MAGIC * Footer formulas referencing the data range are automatically adjusted

# COMMAND ----------

# DBTITLE 1,Install Dependencies
# MAGIC %pip install openpyxl

# COMMAND ----------

# DBTITLE 1,Setup Parameters with Widgets
# Core parameters
dbutils.widgets.text("template_path", "", "Template XLSX path (UC Volume)")
dbutils.widgets.text("output_path", "", "Output directory (UC Volume)")
dbutils.widgets.text("output_filename", "report.xlsx", "Output Filename")

# Table source — one widget per table marker in your template
dbutils.widgets.text("source_sales", "", "Source for {{table:sales}} (SQL or table name)")

# Free-form text fields — one widget per text marker in your template
dbutils.widgets.text("report_title", "", "Report Title")
dbutils.widgets.text("description", "", "Description")
dbutils.widgets.text("generated_date", "", "Generated Date")

# COMMAND ----------

# DBTITLE 1,Core Function: fill_xlsx_template
import copy
import re
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Regex to find all markers: {{name}} or {{table:name}}
MARKER_RE = re.compile(r'\{\{(table:)?(\w+)\}\}')
# Regex to find cell range refs in formulas like SUM(C6:C6)
FORMULA_RANGE_RE = re.compile(r'([A-Z]+)(\d+):([A-Z]+)(\d+)')


def _scan_markers(ws):
    """Scan a worksheet for text and table markers. Returns (text_markers, table_markers)."""
    text_markers = []   # [(cell, marker_name, full_match_string)]
    table_markers = []  # [(cell, table_name)]

    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for match in MARKER_RE.finditer(cell.value):
                    is_table = match.group(1) is not None
                    name = match.group(2)
                    if is_table:
                        table_markers.append((cell, name))
                    else:
                        text_markers.append((cell, name, match.group(0)))

    # Sort table markers by row descending (process bottom-up)
    table_markers.sort(key=lambda x: x[0].row, reverse=True)
    return text_markers, table_markers


def _capture_row_styles(ws, row_idx, min_col, max_col):
    """Capture styles from a row range for later application."""
    styles = []
    for col in range(min_col, max_col + 1):
        cell = ws.cell(row=row_idx, column=col)
        styles.append({
            'font': copy.copy(cell.font),
            'fill': copy.copy(cell.fill),
            'border': copy.copy(cell.border),
            'alignment': copy.copy(cell.alignment),
            'number_format': cell.number_format,
        })
    return styles


def _apply_styles(cell, style_dict):
    """Apply a captured style dict to a cell."""
    cell.font = copy.copy(style_dict['font'])
    cell.fill = copy.copy(style_dict['fill'])
    cell.border = copy.copy(style_dict['border'])
    cell.alignment = copy.copy(style_dict['alignment'])
    cell.number_format = style_dict['number_format']


def _adjust_footer_formulas(ws, footer_row, data_start_row, data_end_row, min_col, max_col):
    """Rewrite footer formulas so range references span the actual data rows."""
    for col in range(min_col, max_col + 1):
        cell = ws.cell(row=footer_row, column=col)
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            def replace_range(m):
                col_start, _, col_end, _ = m.group(1), m.group(2), m.group(3), m.group(4)
                return f'{col_start}{data_start_row}:{col_end}{data_end_row}'
            cell.value = FORMULA_RANGE_RE.sub(replace_range, cell.value)


def fill_xlsx_template(
    template_path: str,
    output_path: str,
    sources: dict,
    text_fields: dict,
    output_filename: str = "report.xlsx"
) -> dict:
    """
    Fill an XLSX template with query results and text, preserving formatting.

    Parameters
    ----------
    template_path : str
        UC Volume path to template XLSX (e.g. /Volumes/catalog/schema/vol/template.xlsx)
    output_path : str
        UC Volume directory for output (e.g. /Volumes/catalog/schema/vol/reports/)
    sources : dict
        Mapping of table marker name to SQL query or table name.
        e.g. {"sales": "SELECT * FROM main.default.sales"}
    text_fields : dict
        Mapping of text marker name to replacement value.
        e.g. {"report_title": "Q1 Sales", "description": "..."}
    output_filename : str
        Output file name.

    Returns
    -------
    dict : Summary of the export.
    """

    # -- Step 1: Copy template to local temp --
    local_template = '/tmp/_xlsx_template_input.xlsx'
    local_output = '/tmp/_xlsx_template_output.xlsx'
    dbutils.fs.cp(template_path, f'file:{local_template}')

    # -- Step 2: Load workbook --
    wb = load_workbook(local_template)

    tables_filled = {}

    # -- Step 3: Process each sheet --
    for ws in wb.worksheets:
        text_markers, table_markers = _scan_markers(ws)

        # -- Step 4: Replace text markers --
        for cell, marker_name, full_match in text_markers:
            if marker_name in text_fields:
                replacement = text_fields[marker_name]
                if cell.value == full_match:
                    # Entire cell is the marker
                    cell.value = replacement
                else:
                    # Marker is embedded in other text
                    cell.value = cell.value.replace(full_match, str(replacement))

        # -- Step 5: Process table markers (bottom-up) --
        for cell, table_name in table_markers:
            if table_name not in sources or not sources[table_name]:
                cell.value = f'(no source for {table_name})'
                continue

            source = sources[table_name]

            # Run query
            if source.strip().upper().startswith(('SELECT', 'WITH')):
                df = spark.sql(source)
            else:
                df = spark.table(source)

            data = df.collect()
            columns = df.columns
            n_rows = len(data)

            template_row = cell.row
            # Determine data column range from template row
            min_col = cell.column
            max_col = min_col + len(columns) - 1
            footer_row = template_row + 1

            # Capture template data row styles
            row_styles = _capture_row_styles(ws, template_row, min_col, max_col)
            template_height = ws.row_dimensions[template_row].height

            # Capture footer row styles before any insertion
            footer_styles = _capture_row_styles(ws, footer_row, min_col, max_col)
            footer_values = []
            for col in range(min_col, max_col + 1):
                footer_values.append(ws.cell(row=footer_row, column=col).value)
            footer_height = ws.row_dimensions[footer_row].height

            if n_rows == 0:
                # No data — clear marker, leave message
                cell.value = 'No data available'
                tables_filled[table_name] = 0
                continue

            # Insert extra rows if needed (we already have 1 row — the template row)
            if n_rows > 1:
                ws.insert_rows(template_row + 1, amount=n_rows - 1)

            # Write data rows with formatting
            for i, row_data in enumerate(data):
                target_row = template_row + i
                if template_height:
                    ws.row_dimensions[target_row].height = template_height
                for j, col_name in enumerate(columns):
                    col_idx = min_col + j
                    value = row_data[col_name]
                    target_cell = ws.cell(row=target_row, column=col_idx)
                    target_cell.value = value if value is not None else ''
                    if j < len(row_styles):
                        _apply_styles(target_cell, row_styles[j])

            # Restore footer row (it was shifted by insert_rows)
            new_footer_row = template_row + n_rows
            if footer_height:
                ws.row_dimensions[new_footer_row].height = footer_height
            for j in range(max_col - min_col + 1):
                col_idx = min_col + j
                fc = ws.cell(row=new_footer_row, column=col_idx)
                fc.value = footer_values[j]
                _apply_styles(fc, footer_styles[j])

            # Adjust footer formulas to span actual data range
            data_start_row = template_row
            data_end_row = template_row + n_rows - 1
            _adjust_footer_formulas(ws, new_footer_row, data_start_row, data_end_row, min_col, max_col)

            tables_filled[table_name] = n_rows

    # -- Step 6: Save output --
    wb.save(local_output)
    full_output = f"{output_path.rstrip('/')}/{output_filename}"
    dbutils.fs.cp(f'file:{local_output}', full_output)

    # Cleanup temp files
    for f in [local_template, local_output]:
        if os.path.exists(f):
            os.remove(f)

    return {
        'template': template_path,
        'output': full_output,
        'text_fields_applied': list(text_fields.keys()),
        'tables_filled': tables_filled,
    }

# COMMAND ----------

# DBTITLE 1,Execute Template Fill
# Get widget values
template_path = dbutils.widgets.get("template_path")
output_path = dbutils.widgets.get("output_path")
output_filename = dbutils.widgets.get("output_filename")

# Validate required parameters
if not template_path:
    raise ValueError("template_path is required")
if not output_path:
    raise ValueError("output_path is required")

# Build sources dict from source_* widgets
# Add entries here for each table marker in your template
sources = {}
source_sales = dbutils.widgets.get("source_sales")
if source_sales:
    sources["sales"] = source_sales

# Build text_fields dict from text widgets
text_fields = {}
for field_name in ["report_title", "description", "generated_date"]:
    val = dbutils.widgets.get(field_name)
    if val:
        text_fields[field_name] = val

# Execute
print(f"Template: {template_path}")
print(f"Output:   {output_path}/{output_filename}")
print(f"Sources:  {list(sources.keys())}")
print(f"Text fields: {list(text_fields.keys())}")
print("-" * 80)

summary = fill_xlsx_template(
    template_path=template_path,
    output_path=output_path,
    sources=sources,
    text_fields=text_fields,
    output_filename=output_filename,
)

print("\nExport completed successfully!")
print("\nSummary:")
for key, value in summary.items():
    print(f"  {key}: {value}")