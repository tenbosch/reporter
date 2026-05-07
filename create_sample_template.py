"""
Generate sample XLSX template and filled report for excel_template_reporter development.

Produces two files:
  - sample_template.xlsx  — template with {{markers}} ready for the reporter notebook
  - sample_report_filled.xlsx — same template filled with synthetic sales data

Run: python create_sample_template.py
"""

import copy
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter

# -- Color palette --
DARK_BLUE = "1F4E79"
MEDIUM_BLUE = "2E75B6"
LIGHT_BLUE = "D6E4F0"
LIGHT_GRAY = "F2F2F2"
WHITE = "FFFFFF"
BLACK = "000000"
DARK_GRAY = "808080"

# -- Reusable styles --
thin_border = Border(
    left=Side(style="thin", color=BLACK),
    right=Side(style="thin", color=BLACK),
    top=Side(style="thin", color=BLACK),
    bottom=Side(style="thin", color=BLACK),
)

header_font = Font(name="Calibri", bold=True, size=11, color=WHITE)
header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")

data_font = Font(name="Calibri", size=11, color=BLACK)
data_fill_even = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
data_fill_odd = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

footer_font = Font(name="Calibri", bold=True, size=11, color=BLACK)
footer_border = Border(
    left=Side(style="thin", color=BLACK),
    right=Side(style="thin", color=BLACK),
    top=Side(style="double", color=BLACK),
    bottom=Side(style="thin", color=BLACK),
)

COLUMNS = ["Product", "Region", "Quantity", "Unit Price", "Total Amount"]
COL_WIDTHS = [22, 16, 12, 14, 16]


def build_template_sheet(ws):
    """Build template structure with markers on a worksheet."""

    # -- Row 1: Title (merged) --
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = "{{report_title}}"
    title_cell.font = Font(name="Calibri", bold=True, size=16, color=WHITE)
    title_cell.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # -- Row 2: Description (merged) --
    ws.merge_cells("A2:E2")
    desc_cell = ws["A2"]
    desc_cell.value = "{{description}}"
    desc_cell.font = Font(name="Calibri", italic=True, size=11, color=BLACK)
    desc_cell.fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
    desc_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    # -- Row 3: Generated date (merged, right-aligned) --
    ws.merge_cells("A3:E3")
    date_cell = ws["A3"]
    date_cell.value = "Generated: {{generated_date}}"
    date_cell.font = Font(name="Calibri", size=9, color=DARK_GRAY)
    date_cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[3].height = 20

    # -- Row 4: Spacer --
    ws.row_dimensions[4].height = 10

    # -- Row 5: Table headers --
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=5, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 24

    # -- Row 6: Template data row with marker --
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=6, column=col_idx)
        cell.font = data_font
        cell.fill = data_fill_even
        cell.border = thin_border
        if col_idx == 1:
            cell.value = "{{table:sales}}"
            cell.alignment = Alignment(horizontal="left")
        elif col_idx == 3:
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")
        elif col_idx in (4, 5):
            cell.number_format = '$#,##0.00'
            cell.alignment = Alignment(horizontal="right")
        else:
            cell.alignment = Alignment(horizontal="left")
    ws.row_dimensions[6].height = 20

    # -- Row 7: Footer row --
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=7, column=col_idx)
        cell.font = footer_font
        cell.border = footer_border
        if col_idx == 1:
            cell.value = "Total"
        elif col_idx == 3:
            cell.value = "=SUM(C6:C6)"
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")
        elif col_idx == 5:
            cell.value = "=SUM(E6:E6)"
            cell.number_format = '$#,##0.00'
            cell.alignment = Alignment(horizontal="right")

    # -- Column widths --
    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def create_template(output_path="sample_template.xlsx"):
    """Create the template XLSX with markers."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    build_template_sheet(ws)
    wb.save(output_path)
    print(f"Template saved: {output_path}")
    return wb


# -- Synthetic data --
PRODUCTS = [
    "Wireless Keyboard", "USB-C Hub", "Monitor Stand", "Laptop Sleeve",
    "Mechanical Keyboard", "Webcam HD", "Desk Lamp", "Mouse Pad XL",
    "Portable SSD 1TB", "Bluetooth Speaker", "Noise-Cancel Headphones",
    "Cable Organizer", "Phone Stand", "HDMI Cable 6ft", "Screen Protector",
    "Power Strip", "Ergonomic Chair Pad", "Desk Shelf", "Ring Light", "Stylus Pen",
]
REGIONS = ["North", "South", "East", "West", "Central"]


def generate_synthetic_data(n_rows=20):
    """Generate n_rows of synthetic sales data."""
    random.seed(42)
    rows = []
    for _ in range(n_rows):
        product = random.choice(PRODUCTS)
        region = random.choice(REGIONS)
        qty = random.randint(5, 500)
        unit_price = round(random.uniform(9.99, 299.99), 2)
        total = round(qty * unit_price, 2)
        rows.append((product, region, qty, unit_price, total))
    return rows


def create_filled_report(output_path="sample_report_filled.xlsx", n_rows=20):
    """Create a filled report from the template with synthetic data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    build_template_sheet(ws)

    data = generate_synthetic_data(n_rows)

    # Replace text markers
    ws["A1"].value = "Quarterly Sales Report"
    ws["A2"].value = "This report summarizes product sales across all regions for Q1 2026. Data is sourced from the internal sales tracking system."
    ws["A3"].value = "Generated: 2026-04-05"

    # Insert data rows (need n_rows - 1 extra rows since row 6 already exists)
    if len(data) > 1:
        ws.insert_rows(7, amount=len(data) - 1)

    # Capture template row styles from row 6
    template_styles = []
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=6, column=col_idx)
        template_styles.append({
            "font": copy.copy(cell.font),
            "fill_even": copy.copy(data_fill_even),
            "fill_odd": copy.copy(data_fill_odd),
            "border": copy.copy(cell.border),
            "alignment": copy.copy(cell.alignment),
            "number_format": cell.number_format,
        })

    # Write data rows with formatting
    for row_idx, row_data in enumerate(data):
        target_row = 6 + row_idx
        is_even = row_idx % 2 == 0
        for col_idx, value in enumerate(row_data):
            cell = ws.cell(row=target_row, column=col_idx + 1, value=value)
            style = template_styles[col_idx]
            cell.font = copy.copy(style["font"])
            cell.fill = copy.copy(style["fill_even"] if is_even else style["fill_odd"])
            cell.border = copy.copy(style["border"])
            cell.alignment = copy.copy(style["alignment"])
            cell.number_format = style["number_format"]

    # Fix footer formulas — footer is now at row 6 + n_rows
    footer_row = 6 + len(data)
    last_data_row = 5 + len(data)
    ws.cell(row=footer_row, column=3).value = f"=SUM(C6:C{last_data_row})"
    ws.cell(row=footer_row, column=5).value = f"=SUM(E6:E{last_data_row})"

    wb.save(output_path)
    print(f"Filled report saved: {output_path}")


if __name__ == "__main__":
    create_template("sample_template.xlsx")
    create_filled_report("sample_report_filled.xlsx", n_rows=20)
    print("Done.")
